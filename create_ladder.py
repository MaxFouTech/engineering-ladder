from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

wb = Workbook()

# Colors
DARK_BLUE = "1F3864"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_YELLOW = "FFF2CC"
LIGHT_GREEN = "E2EFDA"
LIGHT_RED = "FCE4EC"
LIGHT_GRAY = "F2F2F2"
HEADER_YELLOW = "FFFFCC"
HEADER_GREEN = "CCFFCC"
HEADER_RED = "FFCCCC"

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_cell(cell, font=None, fill=None, alignment=None, border=None):
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    if border: cell.border = border

def build_rung_sheet(ws, rung_num, title, subtitle, salary, question, behaviors, promotion_signals, sheet_index):
    ws.sheet_properties.tabColor = DARK_BLUE

    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 35

    row = 1
    # Engineer name row
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    style_cell(c, Font(bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=MED_BLUE),
               Alignment(horizontal='left', vertical='center'))
    c.value = "Engineer Name:"
    ws.row_dimensions[row].height = 30
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        if col > 1:
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)

    row = 2
    # Date row
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    style_cell(c, Font(bold=True, size=11, color=WHITE), PatternFill('solid', fgColor=MED_BLUE),
               Alignment(horizontal='left', vertical='center'))
    c.value = "Review Date:"
    ws.row_dimensions[row].height = 25
    for col in range(1, 7):
        ws.cell(row=row, column=col).border = thin_border
        if col > 1:
            ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)

    row = 3
    # Title row
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    style_cell(c, Font(bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_BLUE),
               Alignment(horizontal='left', vertical='center'))
    c.value = f"Rung {rung_num}: {title}   {subtitle}   {salary}"
    ws.row_dimensions[row].height = 35
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=DARK_BLUE)
        ws.cell(row=row, column=col).border = thin_border

    row = 4
    # Question row
    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    style_cell(c, Font(italic=True, size=11, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_BLUE),
               Alignment(horizontal='left', vertical='center', wrap_text=True))
    c.value = f"Their question: {question}"
    ws.row_dimensions[row].height = 28
    for col in range(1, 7):
        ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
        ws.cell(row=row, column=col).border = thin_border

    row = 5
    # Header row
    headers = ["Observable Behavior", "1\nNot Yet", "2\nEmerging", "3\nConsistent", "4\nExemplary", "Manager's Notes"]
    header_fills = [
        PatternFill('solid', fgColor=DARK_BLUE),
        PatternFill('solid', fgColor=HEADER_RED),
        PatternFill('solid', fgColor=HEADER_YELLOW),
        PatternFill('solid', fgColor=LIGHT_BLUE),
        PatternFill('solid', fgColor=HEADER_GREEN),
        PatternFill('solid', fgColor=DARK_BLUE),
    ]
    header_fonts = [
        Font(bold=True, size=11, color=WHITE),
        Font(bold=True, size=10, color="000000"),
        Font(bold=True, size=10, color="000000"),
        Font(bold=True, size=10, color="000000"),
        Font(bold=True, size=10, color="000000"),
        Font(bold=True, size=11, color=WHITE),
    ]
    ws.row_dimensions[row].height = 35
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        style_cell(c, header_fonts[i], header_fills[i],
                   Alignment(horizontal='center', vertical='center', wrap_text=True), thin_border)

    # Behavior rows
    behavior_start = row + 1
    for idx, behavior in enumerate(behaviors):
        r = behavior_start + idx
        ws.row_dimensions[r].height = 50
        c = ws.cell(row=r, column=1, value=behavior)
        style_cell(c, Font(size=10), PatternFill('solid', fgColor=WHITE if idx % 2 == 0 else LIGHT_GRAY),
                   Alignment(vertical='center', wrap_text=True), thin_border)
        bg = WHITE if idx % 2 == 0 else LIGHT_GRAY
        for col in range(2, 6):
            cell = ws.cell(row=r, column=col)
            style_cell(cell, Font(size=10), PatternFill('solid', fgColor=bg),
                       Alignment(horizontal='center', vertical='center'), thin_border)
            # Data validation: user enters 1 in the chosen column
        # Notes column
        cell = ws.cell(row=r, column=6)
        style_cell(cell, Font(size=10), PatternFill('solid', fgColor=bg),
                   Alignment(vertical='center', wrap_text=True), thin_border)

    behavior_end = behavior_start + len(behaviors) - 1

    # Score row
    score_row = behavior_end + 1
    ws.row_dimensions[score_row].height = 30
    c = ws.cell(row=score_row, column=1, value="TOTAL SCORE")
    style_cell(c, Font(bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=DARK_BLUE),
               Alignment(horizontal='left', vertical='center'), thin_border)

    max_score = len(behaviors) * 4
    # Merged score display across B-E
    ws.merge_cells(f'B{score_row}:E{score_row}')
    # Formula: sum of (col_value * score) for each behavior row
    # Each behavior: user puts any value in one of columns B-E. Score = column index (B=1, C=2, D=3, E=4)
    # We'll use: for each row, IF(E not blank, 4, IF(D not blank, 3, IF(C not blank, 2, IF(B not blank, 1, 0))))
    parts = []
    for r in range(behavior_start, behavior_end + 1):
        parts.append(f'IF(E{r}<>"",4,IF(D{r}<>"",3,IF(C{r}<>"",2,IF(B{r}<>"",1,0))))')
    formula = "=" + "+".join(parts)
    score_cell = ws.cell(row=score_row, column=2, value=formula)
    style_cell(score_cell, Font(bold=True, size=14, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_YELLOW),
               Alignment(horizontal='center', vertical='center'), thin_border)

    # Max score label
    c = ws.cell(row=score_row, column=6, value=f"out of {max_score}")
    style_cell(c, Font(bold=True, size=11, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_YELLOW),
               Alignment(horizontal='left', vertical='center'), thin_border)

    # Percentage row
    pct_row = score_row + 1
    ws.row_dimensions[pct_row].height = 25
    c = ws.cell(row=pct_row, column=1, value="RUNG SCORE %")
    style_cell(c, Font(bold=True, size=11, color=WHITE), PatternFill('solid', fgColor=MED_BLUE),
               Alignment(horizontal='left', vertical='center'), thin_border)
    ws.merge_cells(f'B{pct_row}:E{pct_row}')
    pct_cell = ws.cell(row=pct_row, column=2, value=f'=IF(B{score_row}=0,"",B{score_row}/{max_score})')
    pct_cell.number_format = '0%'
    style_cell(pct_cell, Font(bold=True, size=12, color=MED_BLUE), PatternFill('solid', fgColor=WHITE),
               Alignment(horizontal='center', vertical='center'), thin_border)
    c = ws.cell(row=pct_row, column=6)
    style_cell(c, border=thin_border)

    # Blank row
    blank_row = pct_row + 1

    # Promotion signal row
    promo_row = blank_row + 1
    ws.merge_cells(f'A{promo_row}:F{promo_row}')
    c = ws[f'A{promo_row}']
    style_cell(c, Font(bold=True, size=10, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_BLUE),
               Alignment(horizontal='left', vertical='center', wrap_text=True), thin_border)
    c.value = f"Promotion signal: {promotion_signals}"
    ws.row_dimensions[promo_row].height = 45
    for col in range(1, 7):
        ws.cell(row=promo_row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
        ws.cell(row=promo_row, column=col).border = thin_border

    # Ready to promote row
    ready_row = promo_row + 1
    ws.merge_cells(f'A{ready_row}:F{ready_row}')
    c = ws[f'A{ready_row}']
    style_cell(c, Font(size=11, color=DARK_BLUE), PatternFill('solid', fgColor=WHITE),
               Alignment(horizontal='left', vertical='center'), thin_border)
    c.value = "Ready to promote?   ☐ Not yet   ☐ On track   ☐ Ready now"
    ws.row_dimensions[ready_row].height = 30
    for col in range(1, 7):
        ws.cell(row=ready_row, column=col).border = thin_border

    # 1:1 Notes row
    notes_row = ready_row + 1
    ws.merge_cells(f'A{notes_row}:F{notes_row}')
    c = ws[f'A{notes_row}']
    style_cell(c, Font(bold=True, size=11, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_GRAY),
               Alignment(horizontal='left', vertical='top', wrap_text=True), thin_border)
    c.value = "1:1 Notes:"
    ws.row_dimensions[notes_row].height = 80
    for col in range(1, 7):
        ws.cell(row=notes_row, column=col).fill = PatternFill('solid', fgColor=LIGHT_GRAY)
        ws.cell(row=notes_row, column=col).border = thin_border

    return score_row, pct_row, len(behaviors)

# ---- RUNG DATA ----

rungs = [
    {
        "num": 1, "title": "The Apprentice", "subtitle": "formerly Junior Engineer",
        "salary": "$85K–$120K",
        "question": "Why might this output be wrong?",
        "behaviors": [
            "Flags inconsistencies in code before merging",
            "Asks for context before shipping features",
            "Escalates uncertainty rather than guessing through it",
            "Pull requests include questions, not just solutions",
            "Builds original features and writes real code",
            "Interrogates AI-assisted output — their own and others'",
            "Can explain why AI-generated code will fail in a specific production scenario",
        ],
        "promotion": "Can explain why AI-generated code will fail in production. Has shipped something end-to-end with minimal guidance. Postmortem was more insightful than expected."
    },
    {
        "num": 2, "title": "The Builder", "subtitle": "formerly Mid-Level Engineer",
        "salary": "$120K–$155K",
        "question": "What is this feature protecting the company from?",
        "behaviors": [
            "Owns a feature end-to-end: definition of done, edge cases, monitoring, customer impact",
            "Knows when to throw away AI-generated scaffolding and start again",
            "Writes specs or design docs before prompting — thinks before building",
            "Proactively identifies problems that nobody assigned them",
            "Technical decisions reference company goals without being prompted",
            "Understands what's upstream and downstream of the feature they're building",
            "Has mentored an Apprentice visibly and with measurable result",
        ],
        "promotion": "Has resolved an unassigned problem. Has mentored successfully. Can connect a technical decision to a business outcome without prompting. Doesn't need tickets to know what matters."
    },
    {
        "num": 3, "title": "The Architect", "subtitle": "formerly Senior Engineer",
        "salary": "$155K–$210K",
        "question": "What does this decision cost us in six months?",
        "behaviors": [
            "Walks into cross-functional conversations with answers before questions are finished",
            "The C-Suite understands them — translates technical concepts for executives",
            "Technical instincts show up as financial clarity",
            "Translates technical debt into business cost without being asked",
            "Sees downstream consequences of decisions before anyone else",
            "Has improved the output quality of a team they don't manage",
            "Has driven a consequential build-vs-buy decision and can show the math",
        ],
        "promotion": "Has demonstrably improved output quality of a team they don't manage. Has translated a technical risk into a business risk understood by a non-technical executive. Has driven a build-vs-buy decision with clear math."
    },
    {
        "num": 4, "title": "The Multiplier", "subtitle": "formerly Staff Engineer",
        "salary": "$210K–$300K",
        "question": "How does the team make better decisions because of you?",
        "behaviors": [
            "Engineers leave their code reviews smarter than when they arrived",
            "Their standards get adopted without being mandated",
            "Their departure would be felt across teams, not just their own",
            "Defines how AI-generated output gets evaluated, trusted, and deployed",
            "Builds systems that transfer knowledge rather than hoarding it",
            "Output is the quality of everyone else's judgment, not just features",
            "Impact compounds beyond their own individual contribution",
        ],
        "promotion": "Has changed how the engineering organization evaluates an entire category of work — not just one team. The bar is voluntary adoption: people follow the standard because it made their work better, not because they were told to."
    },
    {
        "num": 5, "title": "The Strategist", "subtitle": "formerly Principal Engineer",
        "salary": "$280K–$450K+",
        "question": "Where do we need to be in two years and what does it cost to get there?",
        "behaviors": [
            "In the room for business decisions before technical implications surface",
            "The CEO treats their input as commercial, not just technical",
            "Can quantify the cost of standing still",
            "Build-vs-buy opinions factor in organizational capacity, not just tech preference",
            "Understands the Engineering Efficiency Ratio and how their decisions move it",
            "Thinks about business problems before the business knows it has a question",
            "Can point to decisions they drove that the company still benefits from years later",
        ],
        "promotion": "Can point to two decisions they made that the company is still benefiting from two years later. Not decisions they recommended — decisions they drove."
    },
]

# Build rung sheets
score_info = []  # (sheet_name, score_row, pct_row, num_behaviors)
for i, rung in enumerate(rungs):
    if i == 0:
        ws = wb.active
        ws.title = f"Rung {rung['num']}"
    else:
        ws = wb.create_sheet(f"Rung {rung['num']}")

    sr, pr, nb = build_rung_sheet(ws, rung['num'], rung['title'], rung['subtitle'],
                                   rung['salary'], rung['question'], rung['behaviors'],
                                   rung['promotion'], i)
    score_info.append((ws.title, sr, pr, nb))

# ---- SUMMARY SHEET ----
ws = wb.create_sheet("Summary", 0)
ws.sheet_properties.tabColor = "1F3864"

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 25
ws.column_dimensions['F'].width = 30

# Title
row = 1
ws.merge_cells(f'A{row}:F{row}')
c = ws[f'A{row}']
style_cell(c, Font(bold=True, size=16, color=WHITE), PatternFill('solid', fgColor=DARK_BLUE),
           Alignment(horizontal='center', vertical='center'))
c.value = "Engineering Ladder Assessment — Summary"
ws.row_dimensions[row].height = 40
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=DARK_BLUE)
    ws.cell(row=row, column=col).border = thin_border

row = 2
ws.merge_cells(f'A{row}:F{row}')
c = ws[f'A{row}']
style_cell(c, Font(italic=True, size=10, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_BLUE),
           Alignment(horizontal='center', vertical='center'))
c.value = "Based on \"The CTO's New Engineering Ladder\" by Etienne de Bruin (March 2026)"
ws.row_dimensions[row].height = 25
for col in range(1, 7):
    ws.cell(row=row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    ws.cell(row=row, column=col).border = thin_border

# Engineer info
row = 4
ws.merge_cells(f'A{row}:B{row}')
c = ws[f'A{row}']
style_cell(c, Font(bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=MED_BLUE),
           Alignment(vertical='center'))
c.value = "Engineer Name:"
ws.cell(row=row, column=2).fill = PatternFill('solid', fgColor=MED_BLUE)
ws.merge_cells(f'C{row}:F{row}')
c = ws[f'C{row}']
style_cell(c, Font(size=12), PatternFill('solid', fgColor=LIGHT_YELLOW),
           Alignment(vertical='center'), thin_border)
ws.row_dimensions[row].height = 30
for col in range(1, 7):
    ws.cell(row=row, column=col).border = thin_border

row = 5
ws.merge_cells(f'A{row}:B{row}')
c = ws[f'A{row}']
style_cell(c, Font(bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=MED_BLUE),
           Alignment(vertical='center'))
c.value = "Review Date:"
ws.cell(row=row, column=2).fill = PatternFill('solid', fgColor=MED_BLUE)
ws.merge_cells(f'C{row}:F{row}')
c = ws[f'C{row}']
style_cell(c, Font(size=12), PatternFill('solid', fgColor=LIGHT_YELLOW),
           Alignment(vertical='center'), thin_border)
ws.row_dimensions[row].height = 30
for col in range(1, 7):
    ws.cell(row=row, column=col).border = thin_border

# Scores table
row = 7
headers = ["Rung", "Score", "Max Score", "Percentage", "Salary Range", "Status"]
header_fill = PatternFill('solid', fgColor=DARK_BLUE)
ws.row_dimensions[row].height = 30
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h)
    style_cell(c, Font(bold=True, size=11, color=WHITE), header_fill,
               Alignment(horizontal='center', vertical='center', wrap_text=True), thin_border)

rung_names = [
    "1: The Apprentice", "2: The Builder", "3: The Architect",
    "4: The Multiplier", "5: The Strategist"
]
salaries = ["$85K–$120K", "$120K–$155K", "$155K–$210K", "$210K–$300K", "$280K–$450K+"]

for i, (sheet_name, sr, pr, nb) in enumerate(score_info):
    r = row + 1 + i
    max_s = nb * 4
    bg = WHITE if i % 2 == 0 else LIGHT_GRAY
    fill = PatternFill('solid', fgColor=bg)

    # Rung name
    c = ws.cell(row=r, column=1, value=rung_names[i])
    style_cell(c, Font(bold=True, size=10), fill, Alignment(vertical='center'), thin_border)

    # Score (linked from rung sheet)
    c = ws.cell(row=r, column=2, value=f"='{sheet_name}'!B{sr}")
    style_cell(c, Font(bold=True, size=11), fill, Alignment(horizontal='center', vertical='center'), thin_border)

    # Max score
    c = ws.cell(row=r, column=3, value=max_s)
    style_cell(c, Font(size=10), fill, Alignment(horizontal='center', vertical='center'), thin_border)

    # Percentage
    c = ws.cell(row=r, column=4, value=f'=IF(B{r}=0,"",B{r}/C{r})')
    c.number_format = '0%'
    style_cell(c, Font(bold=True, size=11), fill, Alignment(horizontal='center', vertical='center'), thin_border)

    # Salary
    c = ws.cell(row=r, column=5, value=salaries[i])
    style_cell(c, Font(size=10), fill, Alignment(horizontal='center', vertical='center'), thin_border)

    # Status (conditional text)
    pct_ref = f'D{r}'
    c = ws.cell(row=r, column=6, value=f'=IF({pct_ref}="","Not assessed",IF({pct_ref}>=0.75,"Strong",IF({pct_ref}>=0.5,"Developing","Needs growth")))')
    style_cell(c, Font(size=10), fill, Alignment(horizontal='center', vertical='center'), thin_border)

    ws.row_dimensions[r].height = 28

# Total row
total_row = row + 6
ws.row_dimensions[total_row].height = 32
c = ws.cell(row=total_row, column=1, value="OVERALL TOTAL")
style_cell(c, Font(bold=True, size=12, color=WHITE), PatternFill('solid', fgColor=DARK_BLUE),
           Alignment(vertical='center'), thin_border)
c = ws.cell(row=total_row, column=2, value=f'=SUM(B{row+1}:B{row+5})')
style_cell(c, Font(bold=True, size=14, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_YELLOW),
           Alignment(horizontal='center', vertical='center'), thin_border)
c = ws.cell(row=total_row, column=3, value=f'=SUM(C{row+1}:C{row+5})')
style_cell(c, Font(bold=True, size=12), PatternFill('solid', fgColor=LIGHT_YELLOW),
           Alignment(horizontal='center', vertical='center'), thin_border)
c = ws.cell(row=total_row, column=4, value=f'=IF(B{total_row}=0,"",B{total_row}/C{total_row})')
c.number_format = '0%'
style_cell(c, Font(bold=True, size=14, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_YELLOW),
           Alignment(horizontal='center', vertical='center'), thin_border)
for col in [5, 6]:
    c = ws.cell(row=total_row, column=col)
    style_cell(c, border=thin_border, fill=PatternFill('solid', fgColor=LIGHT_YELLOW))

# Final Rung Level determination
level_row = total_row + 2
ws.merge_cells(f'A{level_row}:B{level_row}')
c = ws[f'A{level_row}']
style_cell(c, Font(bold=True, size=14, color=WHITE), PatternFill('solid', fgColor=DARK_BLUE),
           Alignment(horizontal='left', vertical='center'))
c.value = "FINAL RUNG LEVEL"
ws.cell(row=level_row, column=2).fill = PatternFill('solid', fgColor=DARK_BLUE)
for col in [1, 2]:
    ws.cell(row=level_row, column=col).border = thin_border
ws.row_dimensions[level_row].height = 40

# The logic: highest rung where score >= 75%
ws.merge_cells(f'C{level_row}:F{level_row}')
# Build nested IF: check from rung 5 down
formula = (
    f'=IF(D12="","Not yet assessed",'
    f'IF(D12>=0.75,"Rung 5: The Strategist",'
    f'IF(D11>=0.75,"Rung 4: The Multiplier",'
    f'IF(D10>=0.75,"Rung 3: The Architect",'
    f'IF(D9>=0.75,"Rung 2: The Builder",'
    f'IF(D8>=0.75,"Rung 1: The Apprentice",'
    f'"Below Rung 1 — Focus on Apprentice behaviors"))))))'
)
c = ws[f'C{level_row}']
c.value = formula
style_cell(c, Font(bold=True, size=16, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_GREEN),
           Alignment(horizontal='center', vertical='center'), thin_border)
for col in range(3, 7):
    ws.cell(row=level_row, column=col).border = thin_border
    if col > 3:
        ws.cell(row=level_row, column=col).fill = PatternFill('solid', fgColor=LIGHT_GREEN)

# Instructions
inst_row = level_row + 2
ws.merge_cells(f'A{inst_row}:F{inst_row}')
c = ws[f'A{inst_row}']
style_cell(c, Font(bold=True, size=11, color=DARK_BLUE), PatternFill('solid', fgColor=LIGHT_BLUE),
           Alignment(horizontal='left', vertical='center'))
c.value = "How to Use This Assessment"
ws.row_dimensions[inst_row].height = 25
for col in range(1, 7):
    ws.cell(row=inst_row, column=col).fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    ws.cell(row=inst_row, column=col).border = thin_border

instructions = [
    "1. Fill in each Rung tab by marking the appropriate score column (type any value, e.g. 'X') for each behavior.",
    "2. Score: 1 = Not Yet, 2 = Emerging, 3 = Consistent, 4 = Exemplary. Only mark ONE column per behavior.",
    "3. The Summary tab auto-calculates scores, percentages, and determines the Final Rung Level.",
    "4. Final Rung Level = highest rung where the engineer scores ≥ 75%.",
    "5. Use the Manager's Notes column and 1:1 Notes section for qualitative observations.",
    "6. Review promotion signals at the bottom of each rung to assess readiness for the next level.",
]
for j, inst in enumerate(instructions):
    r = inst_row + 1 + j
    ws.merge_cells(f'A{r}:F{r}')
    c = ws[f'A{r}']
    style_cell(c, Font(size=10), None, Alignment(vertical='center', wrap_text=True), thin_border)
    c.value = inst
    ws.row_dimensions[r].height = 22
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = thin_border

# Print settings
for sheet in wb.sheetnames:
    s = wb[sheet]
    s.page_setup.orientation = 'landscape'
    s.page_setup.fitToWidth = 1
    s.page_setup.fitToHeight = 0
    s.sheet_properties.pageSetUpPr.fitToPage = True

output_path = '/Users/maxfoutech/Documents/engineering-ladder/engineering_ladder_assessment.xlsx'
wb.save(output_path)
print(f"Saved to {output_path}")
